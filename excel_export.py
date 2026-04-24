"""
MODULE 4: EXCEL AUTOMATION
============================
Gaming Arena, LLC — Auto-Generate Professional Excel Reports

WHAT THIS FILE TEACHES YOU:
- openpyxl library (creating and formatting Excel files from Python)
- Professional formatting (fonts, colors, number formats, column widths)
- Writing formulas vs. hardcoded values
- Structuring a multi-sheet workbook
- Automating a deliverable you'd otherwise build by hand

WHY THIS MATTERS:
  Imagine running 10 scenarios, each producing a full 3-statement model.
  In Excel, that's hours of copy-paste. In Python, it's one function call
  that generates a formatted, print-ready workbook in seconds.

INDUSTRY STANDARD: Color coding in finance models
  Blue text  = hardcoded input / assumption
  Black text = formula / calculated value
  Green text = links from other sheets
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter
import pandas as pd
from model_engine import build_full_model
from scenarios import run_scenario_comparison, build_sensitivity_table
from config import ASSUMPTIONS, STARTUP_COSTS, OPEX_BUDGET, calc_startup_category_total


# =============================================================================
# STYLE CONSTANTS
# =============================================================================
# PYTHON CONCEPT: Defining reusable style objects.
# Instead of repeating Font(bold=True, size=12) everywhere, define it once.

# Fonts
FONT_TITLE = Font(name="Arial", bold=True, size=14, color="1F4E79")
FONT_SECTION = Font(name="Arial", bold=True, size=11, color="1F4E79")
FONT_HEADER = Font(name="Arial", bold=True, size=10, color="FFFFFF")
FONT_LABEL = Font(name="Arial", size=10)
FONT_LABEL_BOLD = Font(name="Arial", bold=True, size=10)
FONT_INPUT = Font(name="Arial", size=10, color="0000FF")      # Blue = input
FONT_FORMULA = Font(name="Arial", size=10, color="000000")     # Black = formula
FONT_SUBTOTAL = Font(name="Arial", bold=True, size=10)

# Fills
FILL_HEADER = PatternFill("solid", fgColor="1F4E79")
FILL_SUBTOTAL = PatternFill("solid", fgColor="D9E2F3")
FILL_INPUT = PatternFill("solid", fgColor="FFF2CC")  # Light yellow = assumption
FILL_NONE = PatternFill(fill_type=None)

# Alignment
ALIGN_CENTER = Alignment(horizontal="center", vertical="center")
ALIGN_RIGHT = Alignment(horizontal="right")
ALIGN_LEFT = Alignment(horizontal="left")

# Borders
THIN_BORDER = Border(
    bottom=Side(style="thin", color="000000"),
)
THICK_BORDER = Border(
    top=Side(style="medium", color="000000"),
    bottom=Side(style="double", color="000000"),
)

# Number formats
FMT_ACCOUNTING = '#,##0;(#,##0);"-"'
FMT_CURRENCY = '$#,##0;($#,##0);"-"'
FMT_PERCENT = '0.0%'
FMT_MULTIPLE = '0.00"x"'


# =============================================================================
# HELPER: Write a DataFrame to a sheet with formatting
# =============================================================================

def write_dataframe_to_sheet(
    ws,
    df: pd.DataFrame,
    start_row: int = 1,
    start_col: int = 1,
    title: str = None,
    format_as_currency: bool = True,
):
    """
    Write a pandas DataFrame to an Excel worksheet with professional formatting.

    PYTHON CONCEPT: This is a reusable function. Instead of writing formatting
    code 5 times (once per sheet), we write it once and call it with different data.
    This is the DRY principle: Don't Repeat Yourself.
    """
    row = start_row

    # Title
    if title:
        ws.cell(row=row, column=start_col, value=title).font = FONT_TITLE
        row += 2

    # Headers
    ws.cell(row=row, column=start_col, value="").font = FONT_HEADER
    for j, col_name in enumerate(df.columns):
        cell = ws.cell(row=row, column=start_col + 1 + j, value=col_name)
        cell.font = FONT_HEADER
        cell.fill = FILL_HEADER
        cell.alignment = ALIGN_CENTER
    row += 1

    # Data rows
    for idx_label, data_row in df.iterrows():
        label = str(idx_label).strip()

        # Determine if this is a subtotal / header row
        is_total = any(kw in label.upper() for kw in ["TOTAL", "EBITDA", "EBIT", "GROSS PROFIT", "NET"])
        is_header = label.isupper() and not any(char.isdigit() for char in label)
        is_spacer = all(c == " " or c == "" for c in label) or data_row.isna().all()
        is_check = "Balance Check" in label

        if is_spacer:
            row += 1
            continue

        # Row label
        label_cell = ws.cell(row=row, column=start_col, value=label)
        if is_total:
            label_cell.font = FONT_SUBTOTAL
            label_cell.fill = FILL_SUBTOTAL
        elif is_header:
            label_cell.font = FONT_SECTION
        else:
            label_cell.font = FONT_LABEL

        # Data values
        for j, value in enumerate(data_row):
            cell = ws.cell(row=row, column=start_col + 1 + j)

            if pd.isna(value) or value is None:
                continue

            cell.value = value

            # Formatting based on value type and row context
            if is_total:
                cell.font = FONT_SUBTOTAL
                cell.fill = FILL_SUBTOTAL
                cell.border = THIN_BORDER

            if is_check:
                cell.value = "PASS" if value else "FAIL"
                cell.font = Font(name="Arial", bold=True, size=10,
                                color="008000" if value else "FF0000")
            elif isinstance(value, bool):
                pass
            elif isinstance(value, (int, float)):
                if "Margin" in label or "Rate" in label or "Utilization" in label:
                    cell.number_format = FMT_PERCENT
                elif "DSCR" in label or "Debt-to" in label or "Runway" in label:
                    cell.number_format = FMT_MULTIPLE
                elif format_as_currency:
                    cell.number_format = FMT_CURRENCY

            cell.alignment = ALIGN_RIGHT

        row += 1

    return row  # return the next available row


# =============================================================================
# BUILD THE FULL WORKBOOK
# =============================================================================

def generate_excel_report(output_path: str = "Gaming_Arena_Financial_Model.xlsx"):
    """
    Generate a complete, formatted Excel workbook with multiple tabs.

    Sheets:
      1. Assumptions   — All model inputs (blue font = editable)
      2. Income Statement — 3-year projected P&L
      3. Cash Flow     — 3-year cash flow statement
      4. Balance Sheet  — 3-year balance sheets
      5. Ratios        — Key financial metrics
      6. Scenarios     — Worst / Base / Best comparison
      7. Sensitivity   — EBITDA sensitivity to hours × price

    PYTHON CONCEPT: Building a complex output step by step.
    Each section below creates one sheet. The model data comes from
    model_engine.py (Module 2), and we just format it for presentation.
    """

    wb = Workbook()

    # =========================================================================
    # SHEET 1: ASSUMPTIONS
    # =========================================================================
    ws = wb.active
    ws.title = "Assumptions"
    ws.sheet_properties.tabColor = "1F4E79"

    row = 1
    ws.cell(row=row, column=1, value="Gaming Arena, LLC — Model Assumptions").font = FONT_TITLE
    row += 2

    # Helper to write assumption sections
    def write_assumption(ws, row, label, value, fmt="currency"):
        ws.cell(row=row, column=1, value=label).font = FONT_LABEL
        cell = ws.cell(row=row, column=2, value=value)
        cell.font = FONT_INPUT  # Blue = editable input
        cell.fill = FILL_INPUT  # Yellow highlight = assumption
        if fmt == "currency":
            cell.number_format = FMT_CURRENCY
        elif fmt == "percent":
            cell.number_format = FMT_PERCENT
        elif fmt == "number":
            cell.number_format = '#,##0'
        elif fmt == "multiple":
            cell.number_format = FMT_MULTIPLE
        return row + 1

    a = ASSUMPTIONS

    ws.cell(row=row, column=1, value="Capacity & Operations").font = FONT_SECTION
    row += 1
    row = write_assumption(ws, row, "Total Gaming Devices", a["capacity"]["total_devices"], "number")
    row = write_assumption(ws, row, "PC Stations", a["capacity"]["pc_stations"], "number")
    row = write_assumption(ws, row, "Console Stations", a["capacity"]["console_stations"], "number")
    row = write_assumption(ws, row, "Operating Hours per Day", a["capacity"]["operating_hours_per_day"], "number")
    row = write_assumption(ws, row, "Days per Year", a["capacity"]["days_per_year"], "number")
    row += 1

    ws.cell(row=row, column=1, value="Pricing & Utilization").font = FONT_SECTION
    row += 1
    row = write_assumption(ws, row, "Price per Device-Hour", a["pricing"]["price_per_hour"], "currency")
    row = write_assumption(ws, row, "Base Case Daily Device-Hours", a["utilization"]["base_case_hours"], "number")
    row = write_assumption(ws, row, "Worst Case Daily Device-Hours", a["utilization"]["worst_case_hours"], "number")
    row = write_assumption(ws, row, "Best Case Daily Device-Hours", a["utilization"]["best_case_hours"], "number")
    row += 1

    ws.cell(row=row, column=1, value="Growth Rates").font = FONT_SECTION
    row += 1
    row = write_assumption(ws, row, "Annual Revenue Growth", a["growth"]["annual_revenue_growth"], "percent")
    row = write_assumption(ws, row, "Annual Expense Growth", a["growth"]["annual_expense_growth"], "percent")
    row += 1

    ws.cell(row=row, column=1, value="Debt & Capital Structure").font = FONT_SECTION
    row += 1
    row = write_assumption(ws, row, "Total Project Cost", a["debt"]["total_project_cost"], "currency")
    row = write_assumption(ws, row, "SBA Loan Amount", a["debt"]["loan_amount"], "currency")
    row = write_assumption(ws, row, "Owner Equity", a["debt"]["owner_equity"], "currency")
    row = write_assumption(ws, row, "Interest Rate", a["debt"]["interest_rate"], "percent")
    row = write_assumption(ws, row, "Annual Principal Payment", a["debt"]["annual_principal_payment"], "currency")
    row += 1

    ws.cell(row=row, column=1, value="Other").font = FONT_SECTION
    row += 1
    row = write_assumption(ws, row, "Annual Depreciation", a["depreciation"]["annual_depreciation"], "currency")
    row = write_assumption(ws, row, "F&B Revenue (Year 1)", a["fnb"]["year1_revenue"], "currency")
    row = write_assumption(ws, row, "F&B COGS %", a["fnb"]["cogs_pct"], "percent")
    row = write_assumption(ws, row, "Merchant Processing %", a["fees"]["merchant_processing_pct"], "percent")

    ws.column_dimensions["A"].width = 35
    ws.column_dimensions["B"].width = 18

    # =========================================================================
    # SHEETS 2-5: FINANCIAL STATEMENTS + RATIOS
    # =========================================================================
    model = build_full_model()

    sheets_config = [
        ("Income Statement", model["income_statement"], "1F4E79"),
        ("Cash Flow", model["cash_flow"], "2E75B6"),
        ("Balance Sheet", model["balance_sheet"], "4472C4"),
        ("Ratios", model["ratios"], "5B9BD5"),
    ]

    for sheet_name, df, tab_color in sheets_config:
        ws = wb.create_sheet(title=sheet_name)
        ws.sheet_properties.tabColor = tab_color
        write_dataframe_to_sheet(
            ws, df,
            title=f"Gaming Arena, LLC — {sheet_name}",
            format_as_currency=(sheet_name != "Ratios"),
        )
        ws.column_dimensions["A"].width = 42
        for j in range(len(df.columns)):
            ws.column_dimensions[get_column_letter(j + 2)].width = 18

    # =========================================================================
    # SHEET 6: SCENARIO COMPARISON
    # =========================================================================
    ws = wb.create_sheet(title="Scenarios")
    ws.sheet_properties.tabColor = "FFC000"
    comparison = run_scenario_comparison()
    write_dataframe_to_sheet(
        ws, comparison,
        title="Gaming Arena — Scenario Comparison (Year 1)",
    )
    ws.column_dimensions["A"].width = 30
    for j in range(len(comparison.columns)):
        ws.column_dimensions[get_column_letter(j + 2)].width = 18

    # =========================================================================
    # SHEET 7: SENSITIVITY TABLE
    # =========================================================================
    ws = wb.create_sheet(title="Sensitivity")
    ws.sheet_properties.tabColor = "FF6600"
    sens = build_sensitivity_table("EBITDA")
    write_dataframe_to_sheet(
        ws, sens,
        title="EBITDA Sensitivity — Daily Hours × Price per Hour",
    )
    ws.column_dimensions["A"].width = 16
    for j in range(len(sens.columns)):
        ws.column_dimensions[get_column_letter(j + 2)].width = 14

    # =========================================================================
    # SAVE
    # =========================================================================
    wb.save(output_path)
    print(f"\nExcel report saved: {output_path}")
    print(f"   Sheets: {', '.join(wb.sheetnames)}")
    return output_path


# =============================================================================
# SELF-TEST
# =============================================================================

if __name__ == "__main__":
    print("=" * 70)
    print("GAMING ARENA — EXCEL REPORT GENERATOR")
    print("=" * 70)
    generate_excel_report("Gaming_Arena_Financial_Model.xlsx")
